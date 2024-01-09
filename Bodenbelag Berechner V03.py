import ifcopenshell
import pandas as pd
import os
from tkinter import Tk, filedialog
from datetime import datetime
import openpyxl
import getpass

def extract_space_details(ifc_file):
    ifc_model = ifcopenshell.open(ifc_file)
    spaces = ifc_model.by_type('IfcSpace')
    data = []

    #Extrahiere den Dateinamen aus dem Pfad
    file_name = os.path.basename(ifc_file)

    #Extrahiere den Timestamp aus der IFC-Datei
    timestamp_ifc = None
    owner_history = ifc_model.by_type('IfcOwnerHistory')
    if owner_history:
        timestamp_ifc = owner_history[0].CreationDate
        timestamp_ifc = datetime.utcfromtimestamp(timestamp_ifc).strftime("%Y-%m-%d %H:%M")

    #Exportdatum erstellen
    export_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    for space in spaces:
        net_floor_area = None
        net_perimeter = None
        name = space.Name or ''
        long_name = space.LongName or ''

        for relDefines in space.IsDefinedBy:
            if relDefines.is_a('IfcRelDefinesByProperties'):
                property_set = relDefines.RelatingPropertyDefinition
                if property_set.is_a('IfcElementQuantity'):
                    if property_set.Name == 'BaseQuantities':
                        for quantity in property_set.Quantities:
                            if quantity.Name == 'NetFloorArea':
                                net_floor_area = quantity.AreaValue
                            elif quantity.Name == 'NetPerimeter':
                                net_perimeter = quantity.LengthValue

        data.append([timestamp_ifc, export_timestamp, file_name, name, long_name, net_floor_area, net_perimeter])

    return pd.DataFrame(data, columns=['Timestamp_IFC', 'ExportTimestamp', 'FileName', 'Name', 'LongName', 'NetFloorArea', 'NetPerimeter'])


def main():
    root = Tk()
    root.withdraw()  #Versteckt das Hauptfenster von Tkinter

    file_path = filedialog.askopenfilename(title="Wähle eine IFC-Datei aus")
    if file_path:
        df = extract_space_details(file_path)
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        output_file = os.path.join(desktop, 'Ausmass Bodenbelag.xlsx')

        #Benutzername abrufen
        username = getpass.getuser()  #Dies gibt den aktuellen Benutzernamen zurück

        #Füge hier den Titel der Tabelle und etwaige zusätzliche Überschriften hinzu
        title = "Ausmass Bodenbelag inkl. Sockelleiste"
        subtitle = f"Erstellt durch: {username}, Hochschule Luzern"  # Der Benutzername wird hier eingefügt

        #Erstelle ein ExcelWriter-Objekt und schreibe die Daten ins Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            #Entferne die Spalten aus dem DataFrame, bevor er ins Excel geschrieben wird
            df_export = df.drop(columns=['Timestamp_IFC', 'ExportTimestamp', 'FileName'])
            df_export.to_excel(writer, index=False, startrow=6)

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            #Setze den Titel und Untertitel
            worksheet.merge_cells('A1:E1')
            worksheet['A1'] = title
            worksheet.merge_cells('A2:E2')
            worksheet['A2'] = subtitle

            #Stileinstellungen für den Titel
            title_font = openpyxl.styles.Font(size=16, bold=True)
            worksheet['A1'].font = title_font
            align_left = openpyxl.styles.Alignment(horizontal='left')
            worksheet['A1'].alignment = align_left
            worksheet['A2'].alignment = align_left

            #Erstelle ein Font-Objekt für fette Schrift
            bold_font = openpyxl.styles.Font(bold=True)

            #Füge die einmaligen Informationen unterhalb der Zeile "Erstellt durch" ein und formatiere sie fett
            header_info = ['Timestamp_IFC', 'ExportTimestamp', 'FileName']
            header_values = [df['Timestamp_IFC'].iloc[0], df['ExportTimestamp'].iloc[0], df['FileName'].iloc[0]]

            for col, value in enumerate(header_info, start=1):
                cell = worksheet.cell(row=4, column=col)
                cell.value = value
                cell.font = bold_font  #Setze die Schrift fett
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col, value in enumerate(header_values, start=1):
                cell = worksheet.cell(row=5, column=col)
                cell.value = value
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            #Anpassen der Spaltenbreiten basierend auf dem längsten Inhalt
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells) + 2  # +2 für etwas zusätzlichen Raum
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

        #Die Datei wird automatisch gespeichert
        print(f"Die Details wurden in {output_file} gespeichert und angepasst.")

def as_text(value):
    """Konvertiere einen Wert in Text, auch wenn dieser None ist."""
    if value is None:
        return ""
    return str(value)

if __name__ == '__main__':
    main()
